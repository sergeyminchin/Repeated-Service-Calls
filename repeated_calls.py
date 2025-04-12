import pandas as pd
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# Function to get the absolute path of resources (works with PyInstaller)
def resource_path(relative_path):
    """ Get absolute path to resource, works for development and PyInstaller bundled executable """
    try:
        base_path = sys._MEIPASS  # PyInstaller temp folder
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


# Main class for the application
class RepeatedCallsApp:
    def __init__(self, root):
        print("Initializing the application...")
        self.root = root
        self.root.title("Repeated Calls Analyzer")
        self.root.geometry("500x500")
        self.root.configure(bg="#042c64")

        # Load the icon dynamically
        icon_path = resource_path("politex.ico")
        if os.path.exists(icon_path):
            try:
                self.root.iconbitmap(icon_path)
                print("Icon loaded successfully.")
            except Exception as e:
                print(f"Failed to load icon: {e}")
        else:
            print(f"Icon file not found at {icon_path}")

        self.setup_ui()

    # Setup the UI
    def setup_ui(self):
        print("Setting up UI...")
        # Add a white background for the logo
        logo_frame = tk.Frame(self.root, bg="white")
        logo_frame.pack(pady=10)

        logo_path = resource_path("logo.png")
        if os.path.exists(logo_path):
            try:
                image = Image.open(logo_path)
                image = image.resize((280, 79), Image.LANCZOS)
                self.logo = ImageTk.PhotoImage(image)

                logo_label = tk.Label(logo_frame, image=self.logo, bg="white")
                logo_label.image = self.logo  # Keep a reference to avoid garbage collection
                logo_label.pack()
                print("Logo loaded successfully.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load logo: {e}")

        label = tk.Label(
            self.root,
            text="Upload the Service Calls Report",
            font=("Arial", 16),
            bg="#042c64",
            fg="white"
        )
        label.pack(pady=20)

        # Define consistent button style
        button_style = {
            "bg": "#fc6c04",
            "fg": "white",
            "font": ("Arial", 12),
            "width": 25,
            "height": 2
        }

        upload_btn = tk.Button(self.root, text="Upload File", command=self.upload_file, **button_style)
        upload_btn.pack(pady=10)

        process_btn = tk.Button(self.root, text="Analyze Calls", command=self.process_file, **button_style)
        process_btn.pack(pady=10)

        self.status_label = tk.Label(self.root, text="", bg="#042c64", fg="white", font=("Arial", 10))
        self.status_label.pack(pady=10)

        self.file_path = None

    # Upload file using file dialog
    def upload_file(self):
        print("Uploading file...")
        file_types = [("Excel files", "*.xlsx")]
        self.file_path = filedialog.askopenfilename(title="Select an Excel file", filetypes=file_types)
        if self.file_path:
            print(f"File uploaded: {self.file_path}")
            self.status_label.config(text=f"File uploaded: {os.path.basename(self.file_path)}")

    # Main processing function
    def process_file(self):
        if not self.file_path:
            messagebox.showerror("Error", "Please upload a file first.")
            return

        try:
            output_path = self.choose_save_location()
            if output_path:
                print("Analyzing the file...")
                self.analyze_repeat_calls(self.file_path, output_path)
                messagebox.showinfo("Success", f"Analysis complete. File saved at:\n{output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    # Choose where to save the output file
    def choose_save_location(self):
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Output File As"
        )
        print(f"Output path selected: {save_path}")
        return save_path

    # Analyze repeated calls with dynamic column name handling
    def analyze_repeat_calls(self, file_path, output_file):
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            print("File loaded successfully for analysis.")
        except Exception as e:
            print(f"Error reading the Excel file: {e}")
            messagebox.showerror("Error", f"Failed to read the Excel file: {e}")
            return

        # Handle both possible column names
        call_id_column = None
        if "מס. קריאה" in df.columns:
            call_id_column = "מס. קריאה"
        elif "מספר קריאה" in df.columns:
            call_id_column = "מספר קריאה"
        else:
            messagebox.showerror("Error", "The Excel file must contain either 'מס. קריאה' or 'מספר קריאה' columns.")
            return

        df_relevant = df[["ת. פתיחה", call_id_column, "מס' מכשיר", "לטיפול", "תאור תקלה", "תאור קוד פעולה"]].copy()
        df_relevant["ת. פתיחה"] = pd.to_datetime(df_relevant["ת. פתיחה"], errors="coerce")
        df_relevant = df_relevant.sort_values(by=["מס' מכשיר", "ת. פתיחה"])

        device_calls = defaultdict(list)
        for _, row in df_relevant.iterrows():
            device_id = row["מס' מכשיר"]
            call_id = row[call_id_column]
            open_date = row["ת. פתיחה"]
            technician = row["לטיפול"]
            fault_description = row["תאור תקלה"]
            action_description = row["תאור קוד פעולה"]

            if device_calls[device_id]:
                last_call = device_calls[device_id][-1]
                last_call_date = last_call["ת. פתיחה"]
                if (open_date - last_call_date).days <= 30:
                    last_call["קריאה חוזרת"].append({
                        "קריאה חוזרת": call_id,
                        "ת. פתיחה": open_date,
                        "לטיפול": technician,
                        "תאור תקלה": fault_description,
                        "תאור קוד פעולה": action_description
                    })

            device_calls[device_id].append({
                "קריאה ראשונה": call_id,
                "ת. פתיחה": open_date,
                "לטיפול": technician,
                "מס' מכשיר": device_id,
                "תאור תקלה": fault_description,
                "תאור קוד פעולה": action_description,
                "קריאה חוזרת": []
            })

        technician_data = defaultdict(list)
        total_calls = df_relevant.shape[0]
        total_repeats = 0

        for calls in device_calls.values():
            for call in calls:
                if call["קריאה חוזרת"]:
                    for repeat_call in call["קריאה חוזרת"]:
                        technician_data[call["לטיפול"]].append({
                            "קריאה ראשונה": call["קריאה ראשונה"],
                            "תאור תקלה (קריאה ראשונה)": call["תאור תקלה"],
                            "תאור קוד פעולה (קריאה ראשונה)": call["תאור קוד פעולה"],
                            "קריאה חוזרת": repeat_call["קריאה חוזרת"],
                            "תאור תקלה (קריאה חוזרת)": repeat_call["תאור תקלה"],
                            "תאור קוד פעולה (קריאה חוזרת)": repeat_call["תאור קוד פעולה"],
                            "מס' מכשיר": call["מס' מכשיר"]
                        })
                        total_repeats += 1

        technician_summary = {}
        for tech, records in technician_data.items():
            total_tech_calls = df_relevant[df_relevant["לטיפול"] == tech].shape[0]
            repeat_calls = len(records)
            repeat_call_percentage = (repeat_calls / total_tech_calls) * 100 if total_tech_calls else 0
            technician_summary[tech] = repeat_call_percentage

        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for tech, records in technician_data.items():
                    df_tech = pd.DataFrame(records)
                    df_tech.to_excel(writer, sheet_name=tech, index=False)

                summary_data = {
                    "Total Calls": [total_calls],
                    "Total Repeated Calls": [total_repeats],
                    "Percentage of Repeated Calls": [f"{(total_repeats / total_calls) * 100:.2f}%" if total_calls else "0%"]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name="Summary", index=False)

            # AutoFit columns and add percentages
            self.autofit_and_add_percentages(output_file, technician_summary)
            print("Analysis complete. File saved successfully.")
        except Exception as e:
            print(f"Error saving the Excel file: {e}")
            messagebox.showerror("Error", f"Failed to save the Excel file: {e}")

    # AutoFit columns and add percentage
    def autofit_and_add_percentages(self, file_path, technician_summary):
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Insert percentage for each technician (skip the summary)
            if sheet_name != "Summary":
                percentage = technician_summary.get(sheet_name, 0)
                sheet.insert_rows(1)
                cell = sheet.cell(row=1, column=1)
                cell.value = f"Repeated Calls Percentage: {percentage:.2f}%"
                cell.font = Font(bold=True)

            # AutoFit column widths
            for column_cells in sheet.columns:
                max_length = 0
                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

        workbook.save(file_path)


# Run the application
if __name__ == "__main__":
    print("Starting the app...")
    root = tk.Tk()
    app = RepeatedCallsApp(root)
    root.mainloop()
